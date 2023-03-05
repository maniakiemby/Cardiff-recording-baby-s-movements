import os
from subprocess import Popen
import webbrowser
from datetime import datetime, date, time, timedelta
import threading
from dateutil.relativedelta import relativedelta
import calendar

import kivy
from kivy.config import Config
from kivy.app import App
from kivy.lang import Builder
from kivy.uix.label import Label
from kivy.uix.floatlayout import FloatLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.modalview import ModalView
from kivy.core.window import Window
from kivy.utils import platform
import openpyxl

from myuix import (TitleCurrentDateWidget, ButtonToday, CancelButtonDate, SelectorMonthsWidget, SelectorMonthsLabel,
    SelectorMonthsButtonPrevious, SelectorMonthsButtonNext, CalendarLayoutWidget, CalendarButtonDay
                   )


kivy.require('2.1.0')
__version__ = '0.1'


class Cardiff(GridLayout):
    def __init__(self, **kwargs):
        super(Cardiff, self).__init__(**kwargs)
        self.work_notebook = WorkNotebook()


class WorkNotebook:
    def __init__(self):
        self.path = 'CardiffWorkNotebook.xlsx'
        self.wb_obj = openpyxl.load_workbook(self.path)
        self.sheet_obj = self.wb_obj.active

        self._start_date = None  # pierwsze okienko C4
        self.last_cell = None  # for recovering

    @property
    def _start_date(self):
        return self.__start_date

    @_start_date.setter
    def _start_date(self, value):
        cell_obj = self.sheet_obj.cell(row=30, column=14)
        self.__start_date = cell_obj.value

    def print_cardiff_notebook(self):
        print(platform)
        if platform == 'win':
            os.startfile(self.path)

        if platform == 'linux':
            # f = os.path.join('temp', self.path)
            Popen(['localc', self.path])

        if platform == 'android':
            filepath = os.path.join(os.getcwd(), self.path)
            # filepath = filepath.replace('/', '\\')
            print(os.path.exists('CardiffWorkNotebook.xlsx'))
            print(filepath)
            webbrowser.open_new(f"ms-excel:ofv|u|{self.path}")

        else:
            pass
            # otwórz powiadomienie, bądź znajdź jeszcze inne roziwązanie. (np. wyślij plik emailem)

    def change_value_in_cell(self, row: int, column: int, value: str):
        self.sheet_obj.cell(row=row, column=column, value=value)
        self.wb_obj.save(self.path)

    def which_column(self) -> int:
        delta = datetime.today() - self._start_date

        return delta.days

    @staticmethod
    def which_row() -> int:
        time_now = datetime.now().time()
        delta_time_now = timedelta(hours=time_now.hour, minutes=time_now.minute)
        # the time we start each day
        time_at_begin = time().fromisoformat('09:00')
        delta_time_at_begin = timedelta(hours=time_at_begin.hour, minutes=time_at_begin.minute)
        delta_time_from_begin = delta_time_now - delta_time_at_begin

        # change seconds to minutes and how many half-hour intervals occurred
        return int(delta_time_from_begin.seconds / 60 // 30) + 1

    def get_cell(self, *args):
        row = self.which_row()
        column = self.which_column()
        # table position correction
        row += 3
        column = column + 3
        print(row, column)
        return row, column

    def add_move(self, *args):
        cell = self.get_cell()
        self.last_cell = cell
        row, column = cell
        self.change_value_in_cell(row=row, column=column, value='x')

    def add_move_with_another_hour(self, *args):
        pass

    def remove_move(self, *args):
        if self.last_cell:
            row, column = self.last_cell
            self.change_value_in_cell(row=row, column=column, value='')


class MyApp(App):
    def build(self):
        Window.clearcolor = (151 / 255, 152 / 255, 164 / 255)
        return Cardiff()


if __name__ == '__main__':
    app = MyApp()
    app.run()
